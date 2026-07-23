from __future__ import annotations

import cgi
import hashlib
import hmac
import io
import json
import os
import re
import threading
import time
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime, timezone
from email.utils import format_datetime
from http.server import ThreadingHTTPServer
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import share_app as share

app = share.app
PIPELINE_LOCK = threading.RLock()
PIPELINE_STATE = {"running": False, "message": "", "last_run": ""}


def now_iso() -> str:
    return app.now_iso()


def init_pipeline_schema() -> None:
    app.init_db()
    share.init_share_schema()
    with app.connect_db() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS wps_sources(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                source_type TEXT NOT NULL DEFAULT 'public_url',
                source_url TEXT DEFAULT '',
                drive_id TEXT DEFAULT '',
                file_id TEXT DEFAULT '',
                enabled INTEGER NOT NULL DEFAULT 1,
                interval_minutes INTEGER NOT NULL DEFAULT 60,
                last_hash TEXT DEFAULT '',
                last_scan TEXT,
                last_message TEXT DEFAULT '',
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS wps_records(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source_id INTEGER NOT NULL,
                sheet_name TEXT NOT NULL,
                row_no INTEGER NOT NULL,
                record_key TEXT NOT NULL,
                title TEXT NOT NULL,
                raw_title TEXT DEFAULT '',
                baidu_url TEXT DEFAULT '',
                baidu_passcode TEXT DEFAULT '',
                quark_url TEXT DEFAULT '',
                quark_passcode TEXT DEFAULT '',
                source_score TEXT DEFAULT '',
                source_category TEXT DEFAULT '',
                max_episode INTEGER NOT NULL DEFAULT 0,
                fingerprint TEXT NOT NULL,
                first_seen TEXT NOT NULL,
                last_seen TEXT NOT NULL,
                UNIQUE(source_id, record_key)
            );
            CREATE TABLE IF NOT EXISTS media_library(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                normalized_title TEXT NOT NULL,
                category TEXT DEFAULT '',
                douban_score TEXT DEFAULT '',
                poster_url TEXT DEFAULT '',
                max_episode INTEGER NOT NULL DEFAULT 0,
                update_text TEXT DEFAULT '',
                baidu_task_id INTEGER,
                quark_task_id INTEGER,
                site_resource_id TEXT DEFAULT '',
                publish_status TEXT DEFAULT '',
                publish_error TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(normalized_title)
            );
            """
        )
        defaults = {
            "wps_app_id": "",
            "wps_app_key": "",
            "wps_access_token": "",
            "tmdb_api_key": "",
            "resource_api_url": "",
            "resource_api_token": "",
            "resource_auto_publish": "0",
        }
        for key, value in defaults.items():
            conn.execute(
                "INSERT OR IGNORE INTO settings(key,value) VALUES(?,?)",
                (key, value),
            )
        conn.commit()



def seed_pipeline_settings_from_env() -> None:
    mapping = {
        "WPS_APP_ID": "wps_app_id",
        "WPS_APP_KEY": "wps_app_key",
        "WPS_ACCESS_TOKEN": "wps_access_token",
        "TMDB_API_KEY": "tmdb_api_key",
        "RESOURCE_API_URL": "resource_api_url",
        "RESOURCE_API_TOKEN": "resource_api_token",
    }
    for env_name, setting_name in mapping.items():
        value = os.getenv(env_name, "").strip()
        if value:
            app.set_setting(setting_name, value)


def normalize_title(value: str) -> str:
    text = str(value or "").strip()
    text = re.sub(r"[（(].*?(?:第?\d+季|更新至?\d+|全\d+集).*?[）)]", "", text)
    text = re.sub(r"[\s._\-]*(?:更新至?|更至?|更新|全)?\s*\d+\s*集?.*$", "", text)
    text = re.sub(r"\s+", "", text)
    return text.strip("._-— ") or str(value or "").strip()


def episode_number(*values: str) -> int:
    best = 0
    patterns = [
        r"(?:更新至?|更至?|全)\s*(\d+)\s*集",
        r"1\s*[-~—至]\s*(\d+)\s*集",
        r"第\s*(\d+)\s*集",
        r"(?:EP|E)\s*0*(\d+)\b",
    ]
    for value in values:
        text = str(value or "")
        for pattern in patterns:
            for match in re.findall(pattern, text, flags=re.I):
                try:
                    best = max(best, int(match))
                except ValueError:
                    pass
    return best


def canonical_url(url: str) -> str:
    value = str(url or "").strip().rstrip("，,；;。")
    if not value:
        return ""
    parts = urllib.parse.urlsplit(value)
    query = urllib.parse.parse_qs(parts.query)
    keep = {}
    if query.get("pwd"):
        keep["pwd"] = query["pwd"][0]
    return urllib.parse.urlunsplit(
        (parts.scheme.lower(), parts.netloc.lower(), parts.path.rstrip("/"), urllib.parse.urlencode(keep), "")
    )


def find_header(headers: list[str], aliases: list[str]) -> int | None:
    normalized = [re.sub(r"\s+", "", str(x or "")).lower() for x in headers]
    for alias in aliases:
        a = re.sub(r"\s+", "", alias).lower()
        for idx, value in enumerate(normalized):
            if value == a or a in value:
                return idx
    return None


def rows_from_xlsx(data: bytes) -> list[dict[str, Any]]:
    wb = load_workbook(io.BytesIO(data), data_only=True)
    output: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        values = list(ws.iter_rows(values_only=True))
        if not values:
            continue
        header_row = 0
        for idx, row in enumerate(values[:15]):
            joined = "|".join(str(x or "") for x in row)
            if any(x in joined for x in ("百度", "夸克", "剧名", "名称", "标题")):
                header_row = idx
                break
        headers = [str(x or "").strip() for x in values[header_row]]
        title_i = find_header(headers, ["电视剧名称", "剧名", "名称", "标题", "资源名称", "列表完整标题"])
        baidu_i = find_header(headers, ["百度网盘", "百度链接", "最新百度网盘"])
        quark_i = find_header(headers, ["夸克网盘", "夸克链接", "最新夸克网盘"])
        baidu_code_i = find_header(headers, ["百度提取码", "提取码"])
        quark_code_i = find_header(headers, ["夸克提取码"])
        score_i = find_header(headers, ["豆瓣评分", "评分"])
        category_i = find_header(headers, ["类型", "分类"])
        for row_no, row in enumerate(values[header_row + 1 :], header_row + 2):
            cells = ["" if x is None else str(x).strip() for x in row]
            joined = " ".join(cells)
            urls = app.extract_links(joined)
            baidu = canonical_url(cells[baidu_i]) if baidu_i is not None and baidu_i < len(cells) else ""
            quark = canonical_url(cells[quark_i]) if quark_i is not None and quark_i < len(cells) else ""
            bp = cells[baidu_code_i] if baidu_code_i is not None and baidu_code_i < len(cells) else ""
            qp = cells[quark_code_i] if quark_code_i is not None and quark_code_i < len(cells) else ""
            for item in urls:
                if item["platform"] == "baidu" and not baidu:
                    baidu, bp = canonical_url(item["url"]), item["passcode"]
                if item["platform"] == "quark" and not quark:
                    quark, qp = canonical_url(item["url"]), item["passcode"]
            if not baidu and not quark:
                continue
            raw_title = cells[title_i] if title_i is not None and title_i < len(cells) else ""
            title = normalize_title(raw_title)
            if not title:
                title = f"{ws.title}-第{row_no}行"
            score = cells[score_i] if score_i is not None and score_i < len(cells) else ""
            category = cells[category_i] if category_i is not None and category_i < len(cells) else ""
            record_key = hashlib.sha256(
                f"{ws.title}|{row_no}|{baidu}|{quark}".encode("utf-8")
            ).hexdigest()
            output.append(
                {
                    "sheet_name": ws.title,
                    "row_no": row_no,
                    "record_key": record_key,
                    "title": title,
                    "raw_title": raw_title,
                    "baidu_url": baidu,
                    "baidu_passcode": bp,
                    "quark_url": quark,
                    "quark_passcode": qp,
                    "source_score": score,
                    "source_category": category,
                    "max_episode": episode_number(raw_title, joined),
                }
            )
    return output


def kso_headers(method: str, uri: str, body: str = "") -> dict[str, str]:
    app_id = app.get_setting("wps_app_id")
    app_key = app.get_setting("wps_app_key")
    token = app.get_setting("wps_access_token")
    if not (app_id and app_key and token):
        raise RuntimeError("WPS开放平台模式缺少 APPID、APPKEY 或 access_token")
    content_type = "application/json"
    kso_date = format_datetime(datetime.now(timezone.utc), usegmt=True)
    body_hash = hashlib.sha256(body.encode()).hexdigest() if body else ""
    raw = "KSO-1" + method + uri + content_type + kso_date + body_hash
    signature = hmac.new(app_key.encode(), raw.encode(), hashlib.sha256).hexdigest()
    return {
        "Content-Type": content_type,
        "X-Kso-Date": kso_date,
        "X-Kso-Authorization": f"KSO-1 {app_id}:{signature}",
        "Authorization": f"Bearer {token}",
    }


def fetch_source_bytes(source: dict[str, Any]) -> bytes:
    source_type = source["source_type"]
    if source_type == "wps_openapi":
        uri = f"/v7/drives/{source['drive_id']}/files/{source['file_id']}/download"
        req = urllib.request.Request(
            "https://openapi.wps.cn" + uri,
            headers=kso_headers("GET", uri),
        )
        with urllib.request.urlopen(req, timeout=60) as res:
            payload = json.loads(res.read().decode("utf-8"))
        download_url = str((payload.get("data") or {}).get("url") or "")
        if not download_url:
            raise RuntimeError(payload.get("msg") or "WPS没有返回下载地址")
        with urllib.request.urlopen(download_url, timeout=120) as res:
            return res.read()
    url = str(source["source_url"] or "").strip()
    if not url:
        raise RuntimeError("共享文档地址为空")
    req = urllib.request.Request(
        url,
        headers={"User-Agent": app.QUARK_UA, "Accept": "*/*"},
    )
    with urllib.request.urlopen(req, timeout=120) as res:
        data = res.read()
        content_type = str(res.headers.get("Content-Type") or "")
    if data[:2] == b"PK":
        return data
    if b"<html" in data[:1000].lower() or "text/html" in content_type:
        text = data.decode("utf-8", "ignore")
        links = re.findall(r'https?://[^"\']+\.(?:xlsx|xls)(?:\?[^"\']*)?', text, flags=re.I)
        if not links:
            raise RuntimeError("该WPS分享页未暴露可下载的Excel文件；请改用WPS开放平台模式")
        with urllib.request.urlopen(links[0], timeout=120) as res:
            return res.read()
    return data


def ensure_task(platform: str, title: str, url: str, passcode: str) -> int | None:
    if not url:
        return None
    save_root = app.get_setting("baidu_transfer_dir", "/资源数据").rstrip("/")
    save_path = f"{save_root}/{app.clean_folder_name(title)}"
    with app.connect_db() as conn:
        row = conn.execute(
            "SELECT id FROM tasks WHERE platform=? AND share_url=? ORDER BY id LIMIT 1",
            (platform, url),
        ).fetchone()
        if row:
            task_id = int(row["id"])
            conn.execute(
                "UPDATE tasks SET name=?,passcode=?,enabled=1,auto_update=1 WHERE id=?",
                (title, passcode, task_id),
            )
        else:
            cur = conn.execute(
                """INSERT INTO tasks(
                    platform,name,share_url,passcode,save_path,interval_minutes,
                    enabled,status,last_message,created_at,auto_update
                ) VALUES(?,?,?,?,?,?,?,?,?,?,1)""",
                (
                    platform,
                    title,
                    url,
                    passcode,
                    save_path,
                    60,
                    1,
                    "待扫描",
                    "由WPS共享文档自动导入",
                    now_iso(),
                ),
            )
            task_id = int(cur.lastrowid)
        conn.commit()
    if task_id:
        app.IMPORT_QUEUE.put(task_id)
    return task_id


def tmdb_metadata(title: str) -> tuple[str, str]:
    key = app.get_setting("tmdb_api_key")
    if not key:
        return "", ""
    query = urllib.parse.urlencode({"api_key": key, "language": "zh-CN", "query": title})
    for kind in ("tv", "movie"):
        try:
            with urllib.request.urlopen(
                f"https://api.themoviedb.org/3/search/{kind}?{query}", timeout=30
            ) as res:
                payload = json.loads(res.read().decode("utf-8"))
            results = payload.get("results") or []
            if not results:
                continue
            item = results[0]
            poster = str(item.get("poster_path") or "")
            poster_url = f"https://image.tmdb.org/t/p/w500{poster}" if poster else ""
            genres = item.get("genre_ids") or []
            genre_map = {
                10759: "动作冒险", 16: "动画", 35: "喜剧", 80: "犯罪",
                99: "纪录片", 18: "剧情", 10751: "家庭", 9648: "悬疑",
                10765: "科幻奇幻", 10766: "肥皂剧", 10768: "战争政治",
                37: "西部", 28: "动作", 12: "冒险", 14: "奇幻",
                27: "恐怖", 10749: "爱情", 878: "科幻", 53: "惊悚",
            }
            category = "/".join(genre_map.get(int(x), "") for x in genres[:3])
            return poster_url, category.strip("/")
        except Exception:
            continue
    return "", ""


def upsert_media(record: dict[str, Any], baidu_task: int | None, quark_task: int | None) -> None:
    normalized = normalize_title(record["title"])
    poster, detected_category = tmdb_metadata(normalized)
    category = record.get("source_category") or detected_category
    score = record.get("source_score") or ""
    max_ep = int(record.get("max_episode") or 0)
    update_text = f"更新至{max_ep}" if max_ep else "持续更新"
    with app.connect_db() as conn:
        old = conn.execute(
            "SELECT * FROM media_library WHERE normalized_title=?",
            (normalized,),
        ).fetchone()
        if old:
            new_ep = max(int(old["max_episode"] or 0), max_ep)
            conn.execute(
                """UPDATE media_library SET
                    title=?,category=CASE WHEN ?<>'' THEN ? ELSE category END,
                    douban_score=CASE WHEN ?<>'' THEN ? ELSE douban_score END,
                    poster_url=CASE WHEN ?<>'' THEN ? ELSE poster_url END,
                    max_episode=?,update_text=?,
                    baidu_task_id=COALESCE(?,baidu_task_id),
                    quark_task_id=COALESCE(?,quark_task_id),
                    updated_at=?
                   WHERE id=?""",
                (
                    record["title"], category, category, score, score,
                    poster, poster, new_ep,
                    f"更新至{new_ep}" if new_ep else "持续更新",
                    baidu_task, quark_task, now_iso(), old["id"],
                ),
            )
        else:
            conn.execute(
                """INSERT INTO media_library(
                    title,normalized_title,category,douban_score,poster_url,
                    max_episode,update_text,baidu_task_id,quark_task_id,
                    created_at,updated_at
                ) VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    record["title"], normalized, category, score, poster,
                    max_ep, update_text, baidu_task, quark_task,
                    now_iso(), now_iso(),
                ),
            )
        conn.commit()


def import_records(source_id: int, records: list[dict[str, Any]]) -> dict[str, int]:
    stats = {"records": 0, "new": 0, "changed": 0, "tasks": 0}
    seen = now_iso()
    with app.connect_db() as conn:
        for record in records:
            fingerprint = hashlib.sha256(
                json.dumps(record, ensure_ascii=False, sort_keys=True).encode("utf-8")
            ).hexdigest()
            old = conn.execute(
                "SELECT * FROM wps_records WHERE source_id=? AND record_key=?",
                (source_id, record["record_key"]),
            ).fetchone()
            if old:
                if old["fingerprint"] != fingerprint:
                    stats["changed"] += 1
                conn.execute(
                    """UPDATE wps_records SET title=?,raw_title=?,baidu_url=?,
                       baidu_passcode=?,quark_url=?,quark_passcode=?,
                       source_score=?,source_category=?,max_episode=?,
                       fingerprint=?,last_seen=? WHERE id=?""",
                    (
                        record["title"], record["raw_title"],
                        record["baidu_url"], record["baidu_passcode"],
                        record["quark_url"], record["quark_passcode"],
                        record["source_score"], record["source_category"],
                        record["max_episode"], fingerprint, seen, old["id"],
                    ),
                )
            else:
                stats["new"] += 1
                conn.execute(
                    """INSERT INTO wps_records(
                       source_id,sheet_name,row_no,record_key,title,raw_title,
                       baidu_url,baidu_passcode,quark_url,quark_passcode,
                       source_score,source_category,max_episode,fingerprint,
                       first_seen,last_seen
                    ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (
                        source_id, record["sheet_name"], record["row_no"],
                        record["record_key"], record["title"], record["raw_title"],
                        record["baidu_url"], record["baidu_passcode"],
                        record["quark_url"], record["quark_passcode"],
                        record["source_score"], record["source_category"],
                        record["max_episode"], fingerprint, seen, seen,
                    ),
                )
            stats["records"] += 1
        conn.commit()
    for record in records:
        bt = ensure_task("baidu", record["title"], record["baidu_url"], record["baidu_passcode"])
        qt = ensure_task("quark", record["title"], record["quark_url"], record["quark_passcode"])
        stats["tasks"] += int(bool(bt)) + int(bool(qt))
        upsert_media(record, bt, qt)
    return stats


def scan_source(source_id: int, uploaded: bytes | None = None) -> dict[str, Any]:
    with app.connect_db() as conn:
        row = conn.execute("SELECT * FROM wps_sources WHERE id=?", (source_id,)).fetchone()
    if not row:
        raise RuntimeError("WPS数据源不存在")
    source = dict(row)
    data = uploaded if uploaded is not None else fetch_source_bytes(source)
    digest = hashlib.sha256(data).hexdigest()
    if digest == source.get("last_hash") and uploaded is None:
        with app.connect_db() as conn:
            conn.execute(
                "UPDATE wps_sources SET last_scan=?,last_message=? WHERE id=?",
                (now_iso(), "文件内容未变化", source_id),
            )
            conn.commit()
        return {"records": 0, "new": 0, "changed": 0, "tasks": 0, "unchanged": True}
    records = rows_from_xlsx(data)
    result = import_records(source_id, records)
    with app.connect_db() as conn:
        conn.execute(
            "UPDATE wps_sources SET last_hash=?,last_scan=?,last_message=? WHERE id=?",
            (
                digest,
                now_iso(),
                f"识别{result['records']}行，新增{result['new']}行，变化{result['changed']}行",
                source_id,
            ),
        )
        conn.commit()
    return result


def scan_all_tasks() -> None:
    with app.connect_db() as conn:
        ids = [int(r["id"]) for r in conn.execute("SELECT id FROM tasks WHERE enabled=1 ORDER BY id")]
    for task_id in ids:
        try:
            app.run_operation(task_id, "scan")
            task = app.task_row(task_id)
            if task and int(task["pending_files"] or 0) > 0:
                app.run_operation(task_id, "update")
            task = app.task_row(task_id)
            if task and not task["own_share_url"]:
                share.generate_share(task_id, force=False)
        except Exception as exc:
            app.add_log(task_id, "ERROR", f"流水线扫描失败：{app.friendly_error_message(exc)}")


def source_monitor_loop() -> None:
    while True:
        try:
            with app.connect_db() as conn:
                sources = [
                    dict(r) for r in conn.execute(
                        "SELECT * FROM wps_sources WHERE enabled=1 ORDER BY id"
                    )
                ]
            now = datetime.now()
            for source in sources:
                due = True
                if source.get("last_scan"):
                    try:
                        last = datetime.fromisoformat(str(source["last_scan"]))
                        due = (now - last).total_seconds() >= int(source["interval_minutes"] or 60) * 60
                    except Exception:
                        due = True
                if due:
                    try:
                        scan_source(int(source["id"]))
                    except Exception as exc:
                        with app.connect_db() as conn:
                            conn.execute(
                                "UPDATE wps_sources SET last_scan=?,last_message=? WHERE id=?",
                                (now_iso(), app.friendly_error_message(exc), source["id"]),
                            )
                            conn.commit()
            time.sleep(60)
        except Exception:
            time.sleep(60)


def run_full_pipeline() -> None:
    with PIPELINE_LOCK:
        if PIPELINE_STATE["running"]:
            return
        PIPELINE_STATE["running"] = True
        PIPELINE_STATE["message"] = "正在扫描WPS共享文档"
    try:
        with app.connect_db() as conn:
            source_ids = [int(r["id"]) for r in conn.execute("SELECT id FROM wps_sources WHERE enabled=1")]
        for idx, source_id in enumerate(source_ids, 1):
            PIPELINE_STATE["message"] = f"扫描WPS数据源 {idx}/{len(source_ids)}"
            scan_source(source_id)
        PIPELINE_STATE["message"] = "正在扫描全部网盘任务并转存更新"
        scan_all_tasks()
        PIPELINE_STATE["message"] = "全部同步完成"
        PIPELINE_STATE["last_run"] = now_iso()
    except Exception as exc:
        PIPELINE_STATE["message"] = app.friendly_error_message(exc)
    finally:
        PIPELINE_STATE["running"] = False


def task_share(task_id: int | None) -> dict[str, str]:
    if not task_id:
        return {"url": "", "passcode": "", "status": "无任务"}
    task = app.task_row(int(task_id))
    if not task:
        return {"url": "", "passcode": "", "status": "任务不存在"}
    return {
        "url": str(task["own_share_url"] or ""),
        "passcode": str(task["own_passcode"] or ""),
        "status": str(task["status"] or ""),
    }


def media_rows() -> list[dict[str, Any]]:
    with app.connect_db() as conn:
        rows = [dict(r) for r in conn.execute("SELECT * FROM media_library ORDER BY id DESC")]
    for row in rows:
        row["baidu"] = task_share(row.get("baidu_task_id"))
        row["quark"] = task_share(row.get("quark_task_id"))
    return rows


def export_library() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "电视剧资源库"
    headers = [
        "序号", "封面图地址", "电视剧名称", "豆瓣评分", "类型",
        "百度网盘链接", "百度提取码", "夸克网盘链接", "夸克提取码",
        "更新说明", "转存状态", "最后更新时间",
    ]
    ws.append(headers)
    for col, text in enumerate(headers, 1):
        c = ws.cell(1, col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="4F46E5")
        c.alignment = Alignment(horizontal="center")
    for idx, row in enumerate(reversed(media_rows()), 1):
        status = f"百度:{row['baidu']['status']}；夸克:{row['quark']['status']}"
        ws.append(
            [
                idx, row.get("poster_url", ""), row.get("title", ""),
                row.get("douban_score", ""), row.get("category", ""),
                row["baidu"]["url"], row["baidu"]["passcode"],
                row["quark"]["url"], row["quark"]["passcode"],
                row.get("update_text", ""), status, row.get("updated_at", ""),
            ]
        )
    widths = [8, 42, 28, 12, 24, 52, 14, 52, 14, 18, 38, 22]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def publish_media(media_id: int, progress_only: bool = False) -> dict[str, Any]:
    api_url = app.get_setting("resource_api_url").rstrip("/")
    token = app.get_setting("resource_api_token")
    if not api_url:
        raise RuntimeError("尚未配置资源站API地址")
    with app.connect_db() as conn:
        row = conn.execute("SELECT * FROM media_library WHERE id=?", (media_id,)).fetchone()
    if not row:
        raise RuntimeError("资源不存在")
    item = dict(row)
    baidu = task_share(item.get("baidu_task_id"))
    quark = task_share(item.get("quark_task_id"))
    ep = int(item.get("max_episode") or 0)
    full_title = f"{item['title']}.更新{ep}" if ep else item["title"]
    payload = {
        "external_key": item["normalized_title"],
        "full_title": full_title,
        "poster_title": full_title,
        "update_description": f"更新至{ep}" if ep else "持续更新",
        "category": item.get("category") or "影视",
        "score": item.get("douban_score") or "",
        "poster_url": item.get("poster_url") or "",
        "baidu_url": baidu["url"],
        "baidu_passcode": baidu["passcode"],
        "quark_url": quark["url"],
        "quark_passcode": quark["passcode"],
        "enabled": True,
        "resource_status": "持续更新",
    }
    # 按用户要求：payload 中完全不包含年份。
    if progress_only and item.get("site_resource_id"):
        endpoint = f"{api_url}/api/admin/resources/{item['site_resource_id']}/progress"
        payload = {
            "full_title": full_title,
            "poster_title": full_title,
            "update_description": payload["update_description"],
            "baidu_url": baidu["url"],
            "baidu_passcode": baidu["passcode"],
            "quark_url": quark["url"],
            "quark_passcode": quark["passcode"],
        }
        method = "PATCH"
    else:
        endpoint = f"{api_url}/api/admin/resources/upsert"
        method = "POST"
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    headers = {"Content-Type": "application/json"}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    req = urllib.request.Request(endpoint, data=body, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req, timeout=60) as res:
            result = json.loads(res.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        raise RuntimeError(f"资源站接口返回 HTTP {exc.code}：{exc.read().decode('utf-8','ignore')[:500]}")
    resource_id = str(result.get("id") or (result.get("data") or {}).get("id") or item.get("site_resource_id") or "")
    with app.connect_db() as conn:
        conn.execute(
            "UPDATE media_library SET site_resource_id=?,publish_status=?,publish_error='',updated_at=? WHERE id=?",
            (resource_id, "已发布", now_iso(), media_id),
        )
        conn.commit()
    return result


PIPELINE_PAGE = r"""<!doctype html><html lang="zh-CN"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1"><title>WPS自动同步流水线</title>
<style>
:root{--p:#4f46e5;--ok:#059669;--bg:#f4f6fb;--bd:#e5e7eb}
*{box-sizing:border-box}body{margin:0;background:var(--bg);font-family:-apple-system,BlinkMacSystemFont,"Segoe UI","PingFang SC",sans-serif;color:#182033}
header{padding:25px 5vw;color:white;background:linear-gradient(135deg,#17203a,#4f46e5)}
main{max-width:1380px;margin:20px auto;padding:0 18px}.card{background:#fff;border:1px solid var(--bd);border-radius:16px;padding:18px;margin-bottom:16px}
.grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(250px,1fr));gap:12px}.field{display:flex;flex-direction:column;gap:6px}
input,select{padding:11px;border:1px solid #d8dce7;border-radius:10px;font-size:14px}.btn{border:0;border-radius:10px;padding:10px 14px;cursor:pointer;font-weight:700;text-decoration:none;display:inline-block}.primary{background:var(--p);color:#fff}.success{background:var(--ok);color:#fff}.secondary{background:#eef0f7;color:#30384c}.danger{background:#fee2e2;color:#991b1b}
.actions{display:flex;gap:8px;flex-wrap:wrap;margin-top:12px}table{width:100%;border-collapse:collapse;font-size:13px}th,td{padding:10px 7px;border-bottom:1px solid var(--bd);text-align:left;vertical-align:top}.wrap{overflow:auto}.muted{color:#6b7280}.url{max-width:280px;word-break:break-all}.status{padding:10px;border-radius:10px;background:#eef2ff;margin-top:10px}
</style></head><body><header><h1>WPS共享资源自动流水线</h1><p>监控共享表格、自动转存追更、生成个人分享链接、导出资源库。年份字段已移除。</p></header>
<main>
<div class="card"><h3>添加WPS共享数据源</h3><div class="grid">
<div class="field"><label>名称</label><input id="sname" placeholder="例如：电视剧共享资源表"></div>
<div class="field"><label>读取方式</label><select id="stype"><option value="public_url">公开/直链</option><option value="wps_openapi">WPS开放平台</option></select></div>
<div class="field"><label>共享或下载地址</label><input id="surl" placeholder="公开xlsx下载地址"></div>
<div class="field"><label>Drive ID</label><input id="drive"></div><div class="field"><label>File ID</label><input id="file"></div>
<div class="field"><label>扫描周期（分钟）</label><input id="interval" type="number" value="60" min="5"></div></div>
<div class="actions"><button class="btn primary" onclick="addSource()">保存数据源</button><button class="btn success" onclick="runAll()">一键扫描全部并自动转存</button><a class="btn secondary" href="/api/pipeline/export.xlsx">下载最新资源Excel</a><a class="btn secondary" href="/">返回原控制台</a></div>
<div id="state" class="status"></div></div>

<div class="card"><h3>WPS数据源</h3><div class="wrap"><table><thead><tr><th>名称</th><th>方式</th><th>最后扫描</th><th>结果</th><th>操作</th></tr></thead><tbody id="sources"></tbody></table></div></div>
<div class="card"><h3>影视资源库（无年份）</h3><div class="wrap"><table><thead><tr><th>封面/剧名</th><th>评分/类型</th><th>百度我的链接</th><th>夸克我的链接</th><th>更新</th><th>资源站</th></tr></thead><tbody id="media"></tbody></table></div></div>
</main>
<script>
const esc=s=>String(s??'').replace(/[&<>"']/g,m=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[m]));
async function api(url,opt={}){const r=await fetch(url,{headers:{'Content-Type':'application/json'},...opt});const j=await r.json();if(!j.success)throw new Error(j.message||'操作失败');return j}
async function load(){try{const d=(await api('/api/pipeline')).data;state.innerHTML=`${esc(d.state.message||'等待操作')}　最后完成：${esc(d.state.last_run||'-')}`;
sources.innerHTML=d.sources.length?d.sources.map(x=>`<tr><td><b>${esc(x.name)}</b></td><td>${esc(x.source_type)}</td><td>${esc(x.last_scan||'-')}</td><td>${esc(x.last_message||'-')}</td><td><button class="btn success" onclick="scan(${x.id})">立即扫描</button></td></tr>`).join(''):'<tr><td colspan="5">暂无数据源</td></tr>';
media.innerHTML=d.media.length?d.media.map(x=>`<tr><td>${x.poster_url?`<img src="${esc(x.poster_url)}" style="width:55px;height:76px;object-fit:cover;border-radius:6px">`:''}<br><b>${esc(x.title)}</b></td><td>豆瓣：${esc(x.douban_score||'待确认')}<br>${esc(x.category||'未分类')}</td><td class="url">${x.baidu.url?`<a target="_blank" href="${esc(x.baidu.url)}">${esc(x.baidu.url)}</a>`:'尚未生成'}</td><td class="url">${x.quark.url?`<a target="_blank" href="${esc(x.quark.url)}">${esc(x.quark.url)}</a>`:'尚未生成'}</td><td>${esc(x.update_text)}</td><td>${esc(x.publish_status||'未发布')}<br><button class="btn primary" onclick="publishItem(${x.id},${x.site_resource_id?'true':'false'})">${x.site_resource_id?'只更新标题/说明':'首次发布'}</button></td></tr>`).join(''):'<tr><td colspan="6">暂无资源</td></tr>'}catch(e){state.innerHTML=esc(e.message)}}
async function addSource(){try{await api('/api/pipeline/sources',{method:'POST',body:JSON.stringify({name:sname.value,source_type:stype.value,source_url:surl.value,drive_id:drive.value,file_id:file.value,interval_minutes:Number(interval.value||60)})});load()}catch(e){alert(e.message)}}
async function scan(id){try{alert((await api(`/api/pipeline/sources/${id}/scan`,{method:'POST',body:'{}'})).message);setTimeout(load,1000)}catch(e){alert(e.message)}}
async function runAll(){try{alert((await api('/api/pipeline/run-all',{method:'POST',body:'{}'})).message);setTimeout(load,1000)}catch(e){alert(e.message)}}
async function publishItem(id,progress){try{alert((await api(`/api/pipeline/media/${id}/publish`,{method:'POST',body:JSON.stringify({progress_only:progress})})).message);load()}catch(e){alert(e.message)}}
load();setInterval(load,10000);
</script></body></html>"""


def inject_pipeline_link(page: str) -> str:
    if 'href="/pipeline"' in page:
        return page
    return page.replace(
        "</header>",
        '<div style="margin-top:10px"><a href="/pipeline" style="display:inline-block;background:#fff;color:#3730a3;padding:10px 15px;border-radius:10px;text-decoration:none;font-weight:700">打开WPS自动同步流水线</a></div></header>',
        1,
    )


class PipelineHandler(share.ExtendedHandler):
    server_version = "NetdiskPipeline/5.0"

    def do_GET(self) -> None:
        path = urllib.parse.urlsplit(self.path).path
        if path in ("/", "/index.html"):
            if self.auth_required():
                return
            page = (app.STATIC_DIR / "index.html").read_text(encoding="utf-8")
            page = share.inject_home(page)
            return self.send_bytes(inject_pipeline_link(page).encode(), "text/html; charset=utf-8")
        if path == "/pipeline":
            if self.auth_required():
                return
            return self.send_bytes(PIPELINE_PAGE.encode(), "text/html; charset=utf-8")
        if path == "/api/pipeline":
            if self.auth_required():
                return
            with app.connect_db() as conn:
                sources = [dict(r) for r in conn.execute("SELECT * FROM wps_sources ORDER BY id DESC")]
            return self.send_json({"success": True, "data": {"sources": sources, "media": media_rows(), "state": dict(PIPELINE_STATE)}})
        if path == "/api/pipeline/export.xlsx":
            if self.auth_required():
                return
            return self.send_bytes(
                export_library(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "电视剧总资源库_无年份_最新同步.xlsx",
            )
        return super().do_GET()

    def do_POST(self) -> None:
        path = urllib.parse.urlsplit(self.path).path
        if path == "/api/pipeline/sources":
            if self.auth_required():
                return
            p = self.read_json()
            if not str(p.get("name") or "").strip():
                return self.send_json({"success": False, "message": "请输入数据源名称"}, 400)
            with app.connect_db() as conn:
                cur = conn.execute(
                    """INSERT INTO wps_sources(
                       name,source_type,source_url,drive_id,file_id,
                       interval_minutes,created_at
                    ) VALUES(?,?,?,?,?,?,?)""",
                    (
                        str(p.get("name") or "").strip(),
                        str(p.get("source_type") or "public_url"),
                        str(p.get("source_url") or "").strip(),
                        str(p.get("drive_id") or "").strip(),
                        str(p.get("file_id") or "").strip(),
                        max(5, int(p.get("interval_minutes") or 60)),
                        now_iso(),
                    ),
                )
                source_id = int(cur.lastrowid)
                conn.commit()
            return self.send_json({"success": True, "message": "WPS数据源已保存", "id": source_id})
        m = re.fullmatch(r"/api/pipeline/sources/(\d+)/scan", path)
        if m:
            if self.auth_required():
                return
            source_id = int(m.group(1))
            try:
                result = scan_source(source_id)
                return self.send_json({"success": True, "message": f"扫描完成：新增{result.get('new',0)}，变化{result.get('changed',0)}"})
            except Exception as exc:
                return self.send_json({"success": False, "message": app.friendly_error_message(exc)}, 400)
        if path == "/api/pipeline/run-all":
            if self.auth_required():
                return
            if PIPELINE_STATE["running"]:
                return self.send_json({"success": False, "message": "全量流水线正在运行"}, 409)
            threading.Thread(target=run_full_pipeline, daemon=True).start()
            return self.send_json({"success": True, "message": "已开始扫描WPS、转存更新并生成个人链接"})
        m = re.fullmatch(r"/api/pipeline/media/(\d+)/publish", path)
        if m:
            if self.auth_required():
                return
            payload = self.read_json()
            try:
                publish_media(int(m.group(1)), bool(payload.get("progress_only")))
                return self.send_json({"success": True, "message": "资源站数据已提交"})
            except Exception as exc:
                return self.send_json({"success": False, "message": app.friendly_error_message(exc)}, 400)
        return super().do_POST()


def main() -> None:
    init_pipeline_schema()
    app.seed_settings_from_env()
    seed_pipeline_settings_from_env()
    for worker_no in range(1, app.IMPORT_WORKERS + 1):
        threading.Thread(
            target=app.import_worker_loop,
            args=(worker_no,),
            daemon=True,
        ).start()
    threading.Thread(target=app.scheduler_loop, daemon=True).start()
    threading.Thread(target=source_monitor_loop, daemon=True).start()
    host = os.getenv("HOST", "0.0.0.0")
    port = int(os.getenv("PORT", "8787"))
    server = ThreadingHTTPServer((host, port), PipelineHandler)
    print(f"WPS网盘自动流水线（无年份版）运行在 http://{host}:{port}", flush=True)
    server.serve_forever()


if __name__ == "__main__":
    main()
